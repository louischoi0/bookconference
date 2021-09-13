# -*- encoding: utf-8 -*-
"""
Copyright (c) 2019 - present AppSeed.us
"""

import csv
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.template import loader
from django.http import HttpResponse
from django import template
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.db.utils import IntegrityError
from django.db.models import Max

from django.contrib.auth import authenticate, login

from app.forms import ApplicantForm, NotificationForm, ApplicationForm
from app.models import Applicant, Application, Application
from app.models import Notification

from django.db.models import Count
import datetime

ADMIN_USER = "admin" #kpc5616

def app_detail(request,aid) :
    app = Application.objects.get(aid=aid)
    biz_img = f'images/a{aid}/img_0.jpg'
    pub_img = f'images/a{aid}/img_1.jpg'

    thumb_img = f'images/a{aid}/img_2.jpg'
    detail_img = f'images/a{aid}/img_3.jpg'

    context = { 'app' : app, "biz_img" : biz_img, "pub_img" : pub_img, "thumb_img" : thumb_img, "detail_img": detail_img }    

    if request.user.uid == ADMIN_USER :
        context["is_admin"] = True
    
    html_template = loader.get_template('appdetail.html')
    return HttpResponse(html_template.render(context, request))

def ulogin(request) :
    print(request.POST)
    username = request.POST['username']
    password = request.POST['password']

    try :
        user = Applicant.objects.get(uid=username,password=password)
        login(request, user)
        return redirect(reverse('list_noti'))
    except Applicant.DoesNotExist :
        messages.error(request,"로그인 정보가 올바르지 않습니다.")
        return redirect('/login/')

@login_required(login_url="/login/")
def applications(request) :
    context = {}
    atype = request.GET['atype'] if 'atype' in request.GET else 'all'
    page = request.GET['page'] if 'page' in request.GET else 1
    page = int(page)

    bulk_cnt = 15

    query_string = None if not 'query' in request.GET else request.GET['query']

    request_context = ""
    if 'atype' in request.GET :
        request_context += f"atype={atype}"

    if 'query' in request.GET :
        request_context += f"query={query_string}"
    
    request_context += "&"

    if request.user.uid == ADMIN_USER :
        context["is_admin"] = True

    if request.user.uid == ADMIN_USER:
        apps = Application.objects.all().order_by("-aid")
    else:
        apps = Application.objects.filter(app_user=request.user).order_by("-aid")
    
    if atype is not None and atype != "all":
        apps = apps.filter(app_type=atype)
    
    if query_string is not None :
        _apps = []
        qtype = request.GET['qtype']

        for app in apps :
            target = getattr(app, qtype)
            _query_string = query_string.replace(" ","") 
            _target = target.replace(" ","")
            if _query_string in _target :
                _apps.append(app)

        apps = _apps

    page_cnt = int(len(apps) / bulk_cnt) + 1
    apps = apps[(page-1)*bulk_cnt:page*bulk_cnt]
    
    for app in apps :
        t = datetime.datetime(year=app.created_at.year,month=app.created_at.month,day=app.created_at.day, hour=app.created_at.hour,minute=app.created_at.minute) + datetime.timedelta(hours = 9)
        t = t.strftime("%Y/%m/%d, %H:%M:%S") 
        app.created_at = t 

    context["apps"] = apps
    context["pages"] = list(x + 1 for x in range(page_cnt) )
    context["request_context"] = request_context

    html_template = loader.get_template('applications.html')
    return HttpResponse(html_template.render(context, request))

def joinpage(request) :
    html_template = loader.get_template( 'join.html' )
    context = {}
    return HttpResponse(html_template.render(context, request))

def page_noti(request,postnum):
    noti = Notification.objects.get(pk=postnum)
    html_template = loader.get_template( 'notipage.html' )
    context = { "noti" : noti }
    try :
        if request.user.uid == ADMIN_USER :
            context["is_admin"] = True
    except :
            context["is_admin"] = False

    return HttpResponse(html_template.render(context, request))

@login_required(login_url="/login/")
def list_user(request) :
    users = Applicant.objects.all()    
    is_admin = request.user.uid == 'admin'

    context = { "users" : users , "is_admin" : is_admin }

    html_template = loader.get_template( 'users.html' )
    return HttpResponse(html_template.render(context, request))

def list_noti_temp(request) :
    notis = Notification.objects.all().order_by("-created_at")
    context = {
            "notis" : notis,
    }

    html_template = loader.get_template( 'temp.html' )

def list_noti(request) :
    notis = Notification.objects.all().order_by("-created_at")
    try :
        is_admin = request.user.uid == 'admin'
    except :
        is_admin = False 

    context = {
            "notis" : notis,
            "is_admin" : is_admin
    }

    html_template = loader.get_template( 'notifications.html' )
    return HttpResponse(html_template.render(context, request))
    
def post_noti(request) :
    n = NotificationForm(request.POST)
    n.is_valid()

    n =  Notification(title=n.cleaned_data["title"], content=n.cleaned_data["content"])
    n.save()
    context = {}
    if request.user.uid == ADMIN_USER :
        context["is_admin"] = True
    return redirect(reverse('list_noti'))

def join(request) :
    if request.method != "POST" :
        return HttpResponse("For biden")

    form = ApplicantForm(request.POST)
    form.is_valid()

    form_data = form.cleaned_data
    print(form_data) 
    context = {}
    for k in form_data:
        key = "old_" + k
        context[key] = form_data[k] 
    
    if not "uid" in form_data:
        context["no_uid"] = "(아이디를 입력해 주세요.)"
        context['error_msg'] = context['no_uid'].replace(")","").replace("(","")
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context, request))

    if not "name" in form_data :
        context["no_name"] = "(이름 정보를 입력해 주세요.)"
        context['error_msg'] = context['no_name'].replace(")","").replace("(","")
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context, request))

    if (not "password" in form_data) or (not "password_check" in form_data):
        context["not_eq"] = "(비밀번호를 입력해 주세요.)"
        context['error_msg'] = context['not_eq'].replace(")","").replace("(","")
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context, request))

    if not "contact" in form_data:
        context["no_contact"] = "(연락처를 입력해 주세요.)"
        context['error_msg'] = context['no_contact'].replace(")","").replace("(","")
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context, request))

    if not "email" in form_data:
        context["no_email"] = "(이메일을 입력해 주세요.)"
        context['error_msg'] = context['no_email'].replace(")","").replace("(","")
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context,request))

    if form_data["password"] != form_data["password_check"]:
        context["not_eq"] = "(비밀번호가 동일하지 않습니다.)"
        context['error_msg'] = context['not_eq'].replace(")","").replace("(","")
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context,request))
    
    del form_data["password_check"]
    print(form_data)
    try :
        user = Applicant(**form_data)
        user.save()
    except IntegrityError:
        context["no_uid"] = "(이미 존재하는 아이디입니다.)"
        html_template = loader.get_template( 'join.html' )
        return HttpResponse(html_template.render(context,request))

    login(request, user)
    html_template = loader.get_template( 'notifications.html' )
    return HttpResponse(html_template.render(context,request))

@login_required(login_url="/login/")
def index(request):
    
    context = {}
    context['segment'] = 'index'
    if request.user.uid == ADMIN_USER :
        context["is_admin"] = True

    html_template = loader.get_template( 'index.html' )
    return HttpResponse(html_template.render(context, request))

@login_required(login_url="/login/")
def app_submit(request) :
    context = {}
    if request.user.uid == ADMIN_USER :
        context["is_admin"] = True
    html_template = loader.get_template( 'submit.html' )
    return HttpResponse(html_template.render(context, request))

def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:
        
        load_template      = request.path.split('/')[-1]
        context['segment'] = load_template
        
        html_template = loader.get_template( load_template )
        return HttpResponse(html_template.render(context, request))
        
    except template.TemplateDoesNotExist:

        html_template = loader.get_template( 'page-404.html' )
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template( 'page-500.html' )
        return HttpResponse(html_template.render(context, request))

def handle_uploaded_file(f,aid,index):
    import os 

    try :
        os.mkdir(f"core/static/images/a{aid}")
    except :
        pass

    file_name = f"a{aid}/img_{index}.jpg"

    des = os.path.join("core/static/images/",file_name)
    destination = open(des, 'wb+')

    for chunk in f.chunks(): 
        destination.write(chunk)

    destination.close()
    return file_name


@login_required(login_url="/login/")
def post_submit(request) :
    app_form = ApplicationForm(request.POST) 
    app_form.is_valid()    
    context = {}
    
    app_data = app_form.cleaned_data

    columns = {
    "app_type" : '신청부문',
    "isbn" : "ISBN",
    "book_title" : "도서명",
    "author_name" : "작가명",
    "publisher_name" : "발행처,대표자",
    "inquiries_info" : "담당자명,연락처",
    "published_date" : "초판 1쇄 발행일", 
    "price" : "가격", 
    "book_width" : "가로",
    "book_height": "세로",
    "book_page_cnt" : "페이지",
    "book_sales_cnt" : "판매부수",
    "book_detail" : "도서특징",
    "author_detail" : "작가소개",
    "publisher_detail"  : "출판사 및 대표소개",
    }

    for c in columns :

        if request.POST[c] == "" :
            context = { "noti_msg" : f"{columns[c]} 항목을 입력해 주세요." }
            html_template = loader.get_template('submit.html')
            cached = { x : request.POST[x] for x in columns }
            context['cached'] = cached
            return HttpResponse(html_template.render(context, request))

    kw = { x : request.POST[x] if request.POST[x] else "" for x in columns } 

    tnum = Application.objects.filter(app_type=kw['app_type']).aggregate(Count('aid'))
    c = tnum['aid__count'] + 1
    t = "일반" if kw['app_type'] == 'g' else '번역'
    token = f"{t}-{c}"

    app = Application(app_user=request.user,**kw,app_token=token)
    app.save()

    if 'biz_file'  in request.FILES:
        handle_uploaded_file(request.FILES['biz_file'],app.aid,0)

    if 'pub_file'  in request.FILES:
        handle_uploaded_file(request.FILES['pub_file'],app.aid,1)
    
    if 'thumb_file' in request.FILES:
        handle_uploaded_file(request.FILES['thumb_file'],app.aid,2)

    if 'detail_file' in request.FILES:
        handle_uploaded_file(request.FILES['detail_file'],app.aid,3)

    return redirect(reverse('app_list'))

@login_required(login_url="/login/")
def update_application(request,aid) :
    app_form = ApplicationForm(request.POST) 
    app_form.is_valid()    
    context = {}
    
    app_data = app_form.cleaned_data

    columns = [ 
    "isbn","book_title","author_name",
    "publisher_name","published_date", "price", 
    "book_width","book_height","book_page_cnt",
    "book_sales_cnt","book_detail","author_detail",
    "publisher_detail", "inquiries_info" ]

    kw = { x : request.POST[x] for x in columns } 
    app = Application.objects.get(aid=aid)

    for k in kw :
        v = kw[k]
        setattr(app,k,v)


    app.save()

    if 'biz_file'  in request.FILES:
        handle_uploaded_file(request.FILES['biz_file'],app.aid,0)

    if 'pub_file'  in request.FILES:
        handle_uploaded_file(request.FILES['pub_file'],app.aid,1)
    
    if 'thumb_file' in request.FILES:
        handle_uploaded_file(request.FILES['thumb_file'],app.aid,2)

    if 'detail_file' in request.FILES:
        handle_uploaded_file(request.FILES['detail_file'],app.aid,3)

    return redirect(reverse(f'app_list'))

@login_required(login_url="/login/")
def confirm(request,aid) :
    
    app = Application.objects.get(aid=aid)
    app.accepted_yn = "Y"
    app.save()
     
    return redirect(reverse('app_list'))


@login_required(login_url="/login/")
def delete_noti(request,nid) :

    noti = Notification.objects.get(nid=nid)
    noti.delete()

    return redirect(reverse('list_noti'))

def export(request):
    workbook = Workbook()
    ws = workbook.active

    columns = {
    "app_token" : '접수번호',
    "app_type" : '신청부문',
    "book_title" : "도서명",
    "isbn" : "ISBN",
    "author_name" : "작가명",
    "publisher_name" : "발행처,대표자",
    "inquiries_info" : "담당자명,연락처",
    "published_date" : "초판 1쇄 발행일", 
    "price" : "가격", 
    "book_width" : "가로",
    "book_height": "세로",
    "book_page_cnt" : "페이지",
    "book_sales_cnt" : "판매부수",
    "book_detail" : "도서특징",
    "author_detail" : "작가소개",
    "publisher_detail"  : "출판사 및 대표소개",
    "accepted_yn" : "접수상태"
    }

    # ... worksheet.append(...) all of your data ...
    if request.user.uid == ADMIN_USER :
        apps = Application.objects.all()
    else :
        apps = Application.objects.filter(app_user=request.user)

    for cidx,c in enumerate(columns) :
        ws.cell(1,cidx+1,columns[c])

    for ridx,app in enumerate(apps) :
        for cidx,c in enumerate(columns) :
            v = getattr(app,c)

            if c == "app_type" :
                if v == "g" :
                    ws.cell(ridx+2, cidx+1,"일반")
                else :
                    ws.cell(ridx+2, cidx+1,"번역")

                continue

            ws.cell(ridx+2, cidx+1,v)

    response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=지원내역.xlsx'
    return response